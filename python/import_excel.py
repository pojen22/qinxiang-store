import sqlite3
from openpyxl import load_workbook


def import_excel():

    filename = input("請輸入 Excel 檔名：")

    conn = sqlite3.connect("database/qinxiang_store.db")
    cursor = conn.cursor()

    wb = load_workbook(filename)

    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):

        cursor.execute("""
        UPDATE products
        SET
            barcode=?,
            price=?,
            fridge_max=?,
            fridge_now=?,
            display_order=?,
            restock_threshold=?
        WHERE name=?
        """, (
            row[1],
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[0]
        ))

    conn.commit()

    conn.close()

    print("✅ Excel 匯入完成")